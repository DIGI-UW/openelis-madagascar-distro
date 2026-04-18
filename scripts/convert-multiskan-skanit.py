#!/usr/bin/env python3
"""
Convert Thermo Multiskan FC SkanIt temporary export → well-per-row CSV.

The SkanIt software's "temporary" Excel export is a single-sheet workbook
with a dual plate-grid layout:

    Row 10 header "Abs", 1..12           → absorbance grid starts
    Rows 11..18                          → rows A..H of absorbance values
    Row 20 header "Échantillon", 1..12  → sample ID grid starts
    Rows 21..28                          → rows A..H of sample IDs

(French locale: "Échantillon" = sample; "Blanc" = blank; "NC" = negative
control; "PC" = positive control.)

The generic FileResultParser in the bridge expects a flat well-per-row table
matching the Multiskan profile's column_mapping:

    WellPosition, SampleID, Abs, TestCode

This adapter joins the two grids per well and emits that flat CSV.
Output is semicolon-delimited per the profile's configDefaults (French locale).

Usage:
    python3 convert-multiskan-skanit.py INPUT.xlsx [-o OUTPUT.csv]
    python3 convert-multiskan-skanit.py INPUT.xlsx --test-code "HIV ELISA"
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


ROW_LETTERS = ["A", "B", "C", "D", "E", "F", "G", "H"]


def find_grid_start(worksheet, header_text):
    """Find the row index where column A equals header_text (e.g. 'Abs', 'Échantillon').

    Returns the 1-indexed row number of the header row, or None if not found.
    Checks the first ~50 rows to stay fast.
    """
    for r in range(1, min(50, worksheet.max_row + 1)):
        cell = worksheet.cell(row=r, column=1).value
        if cell is None:
            continue
        if str(cell).strip() == header_text:
            return r
    return None


def read_grid_at(worksheet, header_row):
    """Given the header-row number (containing 'Abs' or 'Échantillon' + 1..12),
    read the next 8 rows (A..H) as an 8x12 plate grid.

    Returns dict { 'A1': value, 'A2': value, ... }.
    """
    wells = {}
    for offset, letter in enumerate(ROW_LETTERS, 1):
        r = header_row + offset
        row_letter_cell = worksheet.cell(row=r, column=1).value
        if row_letter_cell is None or str(row_letter_cell).strip() != letter:
            # Row letter mismatch — either grid ended early or structure changed.
            continue
        for col_idx in range(1, 13):
            # Data cells are columns 2..13 (col B..M), matching 1..12 in header.
            value = worksheet.cell(row=r, column=1 + col_idx).value
            wells[f"{letter}{col_idx}"] = value
    return wells


def classify_sample(sample_id, default_test_code):
    """Return (SampleID, TestCode) for a raw Multiskan sample cell.

    QC conventions: 'Blanc' (blank), 'NC####' (negative), 'PC####' (positive),
    'Échantillon####' or 'S#' (patient).

    Patient IDs are ASCII-normalized (Échantillon → SAMPLE) so downstream
    URL paths and DB accession columns don't need UTF-8-safe handling for
    the É character — only a concern at our demo layer; real sites may
    use different sample conventions.
    """
    if sample_id is None:
        return None, None
    s = str(sample_id).strip()
    if not s:
        return None, None
    s_upper = s.upper()
    if s_upper.startswith("BLANC") or s_upper == "BLANK":
        return s, f"QC_{default_test_code}_BLANK"
    if s_upper.startswith("NC"):
        return s, f"QC_{default_test_code}_NEG"
    if s_upper.startswith("PC"):
        return s, f"QC_{default_test_code}_POS"
    # Patient: ASCII-normalize Échantillon → SAMPLE, keep trailing digits.
    if s.startswith("Échantillon") or s.startswith("Echantillon"):
        digits = s.lstrip("Échantilon").lstrip("Echantilon").lstrip()
        return f"SAMPLE{digits}" if digits else s, default_test_code
    return s, default_test_code


def convert(input_path: Path, output_path: Path, test_code: str):
    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    abs_header_row = find_grid_start(ws, "Abs")
    sample_header_row = find_grid_start(ws, "Échantillon")

    if abs_header_row is None or sample_header_row is None:
        print(
            f"ERROR: could not locate Abs/Échantillon headers "
            f"(Abs@{abs_header_row}, Échantillon@{sample_header_row})",
            file=sys.stderr,
        )
        sys.exit(2)

    abs_grid = read_grid_at(ws, abs_header_row)
    sample_grid = read_grid_at(ws, sample_header_row)

    rows = []
    patient_count = 0
    qc_count = 0
    empty_count = 0

    for row_letter in ROW_LETTERS:
        for col in range(1, 13):
            well = f"{row_letter}{col}"
            sample_raw = sample_grid.get(well)
            abs_value = abs_grid.get(well)

            sample_id, sample_test_code = classify_sample(sample_raw, test_code)
            if sample_id is None or abs_value is None:
                empty_count += 1
                continue

            if sample_test_code.startswith("QC_"):
                qc_count += 1
            else:
                patient_count += 1

            rows.append(
                {
                    "WellPosition": well,
                    "SampleID": sample_id,
                    "Abs": abs_value,
                    "TestCode": sample_test_code,
                }
            )

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["WellPosition", "SampleID", "Abs", "TestCode"],
            delimiter=";",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(
        f"Converted {len(rows)} wells → {output_path}\n"
        f"  Patient: {patient_count}, QC: {qc_count}, Empty: {empty_count}"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Convert Multiskan FC SkanIt xlsx → well-per-row CSV (semicolon-delimited)"
    )
    parser.add_argument("input", type=Path, help="Input xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output CSV path (defaults to INPUT_converted.csv)",
    )
    parser.add_argument(
        "--test-code",
        default="HIV ELISA",
        help="Default test code for patient wells (default: HIV ELISA)",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    output = args.output or args.input.with_name(
        args.input.stem + "_converted.csv"
    )
    convert(args.input, output, args.test_code)


if __name__ == "__main__":
    main()
