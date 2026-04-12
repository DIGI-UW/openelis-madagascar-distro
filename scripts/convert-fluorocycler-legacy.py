#!/usr/bin/env python3
"""
Convert legacy FluoroCycler copy-paste XLSX → standardized FC-XT template.

The FluoroCycler XT has no structured data export. Lab techs manually
copy-paste from FluoroSoftware into Excel, producing a messy 6-column
format with compound strings and concatenated QC notes. This script
converts those legacy files into the standardized 12+1 column template
defined in OGC-420.

Usage:
    python3 convert-fluorocycler-legacy.py HIV-result.xlsx
    python3 convert-fluorocycler-legacy.py HIV-result.xlsx -o FC-XT_HIV-result.xlsx

Input columns (old format):
    Row | Col | Sample ID | Type | Calc. Conc. | Result

Output columns (standardized template):
    SampleID | WellPosition | AssayName | TargetName | TargetNo | CP |
    Interpretation | CalcConc | CalcConcUnit | RunDate | RunID | Notes | Type
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_calc_conc(raw: str | None) -> tuple[float | None, str | None]:
    """Extract numeric concentration and unit from Calc. Conc. field.

    Examples:
        '2.10E6 copies/mL = 6.32 log copies/mL' → (2100000.0, 'copies/mL')
        'Concentration < LOQ'                    → (None, None)
        'Concentration > LOQ'                    → (None, None)
        ''                                       → (None, None)
    """
    if not raw or not raw.strip():
        return None, None
    raw = raw.strip()
    if "< LOQ" in raw or "> LOQ" in raw:
        return None, None

    # Try to extract scientific notation: '2.10E6 copies/mL'
    m = re.match(r"([\d.]+E[+\-]?\d+)\s*(copies/mL|IU/mL)", raw, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1)), m.group(2)
        except ValueError:
            pass

    # Try plain numeric
    m = re.match(r"^([\d.]+)\s*(copies/mL|IU/mL|log copies/mL)?", raw)
    if m:
        try:
            return float(m.group(1)), m.group(2)
        except ValueError:
            pass

    return None, None


def parse_result(raw: str | None) -> tuple[str, float | None, str]:
    """Extract interpretation, CP value, and notes from Result field.

    Examples:
        'HIV-1 +  (CP=26.6)Control(s) failed'
            → ('Detected', 26.6, 'Control(s) failed')
        'HIV-1 - Control(s) failed'
            → ('Not Detected', None, 'Control(s) failed')
        'InvalidControl(s) failed'
            → ('Invalid', None, 'Control(s) failed')
        'Not interpretableControl(s) failed'
            → ('Inconclusive', None, 'Control(s) failed')
        'Negative Control valid'
            → ('Negative Control', None, '')
        'Positive Control invalid'
            → ('Positive Control', None, 'invalid')
    """
    if not raw or not raw.strip():
        return "", None, ""
    raw = raw.strip()

    # Extract CP value if present
    cp = None
    cp_match = re.search(r"\(CP=([\d.]+)\)", raw)
    if cp_match:
        try:
            cp = float(cp_match.group(1))
        except ValueError:
            pass

    # Extract QC note suffix
    notes = ""
    qc_match = re.search(r"(Control\(s\)\s*failed|No Internal Control)", raw, re.IGNORECASE)
    if qc_match:
        notes = qc_match.group(0)

    # Determine interpretation
    interp = ""
    if re.search(r"HIV-1\s*\+", raw):
        interp = "Detected"
    elif re.search(r"HIV-1\s*-", raw):
        interp = "Not Detected"
    elif raw.startswith("Invalid") or "InvalidNo Internal" in raw:
        interp = "Invalid"
    elif raw.startswith("Not interpretable"):
        interp = "Inconclusive"
    elif "Negative Control" in raw:
        interp = "Negative Control"
    elif "Positive Control" in raw:
        interp = "Positive Control"
    elif re.search(r"STD\s+\dE\d", raw):
        interp = "Standard"
    else:
        interp = raw[:50]  # Fallback: truncated raw

    return interp, cp, notes


def convert(input_path: Path, output_path: Path, assay_name: str = "FluoroType HIV-1 VL"):
    """Convert a legacy FluoroCycler XLSX to the standardized template."""
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    # Read headers from row 1
    headers = [cell.value for cell in ws_in[1]]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    required = {"Sample ID", "Type"}
    missing = required - set(col_idx.keys())
    if missing:
        print(f"ERROR: Missing required columns: {missing}", file=sys.stderr)
        print(f"  Found: {list(col_idx.keys())}", file=sys.stderr)
        sys.exit(1)

    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"

    out_headers = [
        "SampleID", "WellPosition", "AssayName", "TargetName", "TargetNo",
        "CP", "Interpretation", "CalcConc", "CalcConcUnit", "RunDate",
        "RunID", "Notes", "Type",
    ]
    for i, h in enumerate(out_headers, 1):
        ws_out.cell(row=1, column=i, value=h)

    row_out = 2
    stats = {"total": 0, "patient": 0, "control": 0, "quantified": 0}

    for row in ws_in.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        if len(vals) < len(headers):
            vals.extend([None] * (len(headers) - len(vals)))

        def get(col_name):
            idx = col_idx.get(col_name)
            if idx is not None and idx < len(vals):
                v = vals[idx]
                return str(v).strip() if v is not None else ""
            return ""

        sample_id = get("Sample ID")
        if not sample_id:
            continue

        row_pos = get("Row")
        col_pos = get("Col")
        well = f"{row_pos}{col_pos}" if row_pos and col_pos else ""
        sample_type = get("Type")
        calc_conc_raw = get("Calc. Conc.")
        result_raw = get("Result")

        # Parse compound fields
        calc_conc, calc_unit = parse_calc_conc(calc_conc_raw)
        interp, cp, notes = parse_result(result_raw)

        # Write output row
        ws_out.cell(row=row_out, column=1, value=sample_id)
        ws_out.cell(row=row_out, column=2, value=well)
        ws_out.cell(row=row_out, column=3, value=assay_name)
        ws_out.cell(row=row_out, column=4, value="HIV-1")
        ws_out.cell(row=row_out, column=5, value=1)
        if cp is not None:
            ws_out.cell(row=row_out, column=6, value=cp)
        ws_out.cell(row=row_out, column=7, value=interp)
        if calc_conc is not None:
            ws_out.cell(row=row_out, column=8, value=calc_conc)
            ws_out.cell(row=row_out, column=9, value=calc_unit)
        elif "< LOQ" in calc_conc_raw:
            ws_out.cell(row=row_out, column=7, value=interp or "Detected")
            ws_out.cell(row=row_out, column=12, value="Below LOQ")
        elif "> LOQ" in calc_conc_raw:
            ws_out.cell(row=row_out, column=12, value="Above LOQ")
        ws_out.cell(row=row_out, column=12, value=notes if notes else None)
        ws_out.cell(row=row_out, column=13, value=sample_type)

        stats["total"] += 1
        if sample_type and sample_type.lower() != "unknown":
            stats["control"] += 1
        else:
            stats["patient"] += 1
        if calc_conc is not None:
            stats["quantified"] += 1

        row_out += 1

    wb_out.save(output_path)
    print(f"Converted {stats['total']} rows → {output_path}")
    print(f"  Patient: {stats['patient']}, Controls: {stats['control']}, Quantified: {stats['quantified']}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert legacy FluoroCycler copy-paste XLSX to standardized template"
    )
    parser.add_argument("input", type=Path, help="Input XLSX (legacy copy-paste format)")
    parser.add_argument("-o", "--output", type=Path, help="Output XLSX (default: FC-XT_<input>)")
    parser.add_argument("--assay", default="FluoroType HIV-1 VL", help="Assay name (default: FluoroType HIV-1 VL)")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: {args.input} not found", file=sys.stderr)
        sys.exit(1)

    output = args.output or args.input.parent / f"FC-XT_{args.input.name}"
    convert(args.input, output, args.assay)


if __name__ == "__main__":
    main()
