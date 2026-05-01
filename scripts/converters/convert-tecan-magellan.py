#!/usr/bin/env python3
"""
Convert Tecan Infinite F50 custom Magellan template → well-per-row CSV.

The lab's Tecan workflow uses a custom Excel workbook (not standard Magellan
ASCII export) with two parallel 96-well grids:

    Plan_plaque sheet    : Row letter A-H × cols 1-12 → Sample IDs + QC labels
    DO_palque  sheet    : Row letter A-H × cols 1-12 → Absorbance (OD_450) values

The generic FileResultParser in the bridge expects a flat well-per-row table
matching the Tecan profile's column_mapping:

    WellPosition, SampleID, OD_450, TestCode

This adapter joins the two sheets per well position and emits that flat CSV.
QC wells are detected by label (NEG/POS) and get a `QC_` TestCode prefix so
the acceptance flow can route them away from patient results.

Output is comma-delimited per the profile's configDefaults.

Usage:
    python3 convert-tecan-magellan.py INPUT.xlsx [-o OUTPUT.csv]
    python3 convert-tecan-magellan.py INPUT.xlsx --test-code "HIV ELISA"

If -o is omitted, writes next to INPUT with _converted.csv suffix.
"""

import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


ROW_LETTERS = ["A", "B", "C", "D", "E", "F", "G", "H"]
# Header is at row 8 on both sheets (col numbers 1..12); data at rows 9..16.
HEADER_ROW = 8
DATA_ROW_START = 9
DATA_ROW_END = 16


def read_plate_grid(worksheet):
    """Return dict { 'A1': value, 'A2': value, ... } from an 8x12 plate grid.

    Assumes the grid is in rows DATA_ROW_START..DATA_ROW_END, with the row
    letter in column B (index 2) and well values in columns 3..14 (1-indexed).
    """
    wells = {}
    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        row_values = [worksheet.cell(row=r, column=c).value for c in range(1, 15)]
        if not row_values or len(row_values) < 2:
            continue
        row_letter = row_values[1]  # col B
        if row_letter not in ROW_LETTERS:
            continue
        # Well values are in columns 3..14 (spreadsheet col C..N) → index 2..13
        for col_idx in range(12):
            value = row_values[2 + col_idx] if 2 + col_idx < len(row_values) else None
            well_position = f"{row_letter}{col_idx + 1}"
            wells[well_position] = value
    return wells


def classify_sample(sample_id, default_test_code):
    """Return (SampleID, TestCode) given the raw sample_id cell.

    QC conventions observed: 'NEG', 'POS ' (trailing space). Patient IDs
    look like 'CG-M4-00-004', 'FE*' (Pareekshak), etc.
    """
    if sample_id is None:
        return None, None
    s = str(sample_id).strip()
    if not s:
        return None, None
    s_upper = s.upper()
    if s_upper == "NEG" or s_upper.startswith("NC"):
        return s, f"QC_{default_test_code}_NEG"
    if s_upper == "POS" or s_upper.startswith("PC"):
        return s, f"QC_{default_test_code}_POS"
    if s_upper.startswith("BLANC") or s_upper == "BLANK":
        return s, f"QC_{default_test_code}_BLANK"
    return s, default_test_code


def convert(input_path: Path, output_path: Path, test_code: str, max_patients: int | None):
    wb = openpyxl.load_workbook(str(input_path), data_only=True)

    # Expected sheet names; fall back by position if renamed.
    sheets = wb.sheetnames
    plan_sheet = "Plan_plaque" if "Plan_plaque" in sheets else sheets[0]
    do_sheet = "DO_palque" if "DO_palque" in sheets else sheets[1]

    plan_grid = read_plate_grid(wb[plan_sheet])
    do_grid = read_plate_grid(wb[do_sheet])

    rows = []
    patient_count = 0
    qc_count = 0
    empty_count = 0

    for row_letter in ROW_LETTERS:
        for col in range(1, 13):
            well = f"{row_letter}{col}"
            sample_raw = plan_grid.get(well)
            od_value = do_grid.get(well)

            sample_id, sample_test_code = classify_sample(sample_raw, test_code)
            if sample_id is None or od_value is None:
                empty_count += 1
                continue

            is_qc = sample_test_code.startswith("QC_")
            if not is_qc and max_patients is not None and patient_count >= max_patients:
                continue

            if is_qc:
                qc_count += 1
            else:
                patient_count += 1

            rows.append(
                {
                    "WellPosition": well,
                    "SampleID": sample_id,
                    "OD_450": od_value,
                    "TestCode": sample_test_code,
                }
            )

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["WellPosition", "SampleID", "OD_450", "TestCode"]
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    print(
        f"Converted {len(rows)} wells → {output_path}\n"
        f"  Patient: {patient_count}, QC: {qc_count}, Empty: {empty_count}"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Convert Tecan Magellan custom xlsx → well-per-row CSV"
    )
    parser.add_argument("input", type=Path, help="Input xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output CSV path (defaults to INPUT_converted.csv)",
    )
    parser.add_argument(
        "--test-code",
        default="HIV ELISA",
        help="Default test code for patient wells (default: HIV ELISA)",
    )
    parser.add_argument(
        "--max-patients",
        type=int,
        default=None,
        help="Limit to first N patient wells (QC always included). "
        "Use for demo fixtures where full 96-well plates exceed test timeouts.",
    )
    args = parser.parse_args()

    if not args.input.exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    output = args.output or args.input.with_name(
        args.input.stem + "_converted.csv"
    )
    convert(args.input, output, args.test_code, args.max_patients)


if __name__ == "__main__":
    main()
